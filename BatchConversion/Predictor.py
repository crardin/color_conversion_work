import re
import openpyxl
from sklearn.neighbors import KNeighborsClassifier


class Predictor:
    __munsellDataFile = "CTAListD50_RoundedLABx.xlsx"
    __conversionData = {}
    __munsellTrainingValues = []
    __labTrainingValues = []
    __predictedMunsellVector = []

    def __init__(self):
        self.neigh = KNeighborsClassifier(n_neighbors=1)
        self.getTrainingData()
        self.trainData()

    def getTrainingData(self):
        wb = openpyxl.load_workbook(self.__munsellDataFile)
        sheet = wb['CTAListD50_RoundedLAB']
        for i in range(2, sheet.max_row + 1):
            munsellValue = self.getMunsellTrainingValue(i, sheet)
            labValue = self.getLabTrainingValue(i, sheet)
            self.__conversionData[labValue] = munsellValue

    def getMunsellTrainingValue(self, i, sheet):
        H1 = sheet.cell(row=i, column=4).value
        munsellValue = str(H1) + " "
        H2 = sheet.cell(row=i, column=5).value
        munsellValue += str(H2) + " "
        V = sheet.cell(row=i, column=6).value
        munsellValue += str(V) + "/"
        C = sheet.cell(row=i, column=7).value
        munsellValue += str(C)
        self.__munsellTrainingValues.append(str(H1) + H2 + " " + str(V) + "/" + str(C))
        return munsellValue

    def getLabTrainingValue(self, i, sheet):
        L = sheet.cell(row=i, column=8).value
        labValue = str(L) + " "
        a = sheet.cell(row=i, column=9).value
        labValue += str(a) + " "
        b = sheet.cell(row=i, column=10).value
        labValue += str(b)
        self.__labTrainingValues.append([L, a, b])
        return labValue

    def trainData(self):
        self.neigh.fit(self.__labTrainingValues, self.__munsellTrainingValues)

    def getMunsellValue(self, inputString):
        inputString.strip()
        splitString = re.findall('(\d+\.?\d|\d+|\w+)', inputString)
        splitString[0] = float(splitString[0])
        splitString[2] = int(splitString[2])
        splitString[3] = int(splitString[3])
        return splitString

    def predictNominalMunsellVector(self, LabVector):
        predictedValue = self.neigh.predict([LabVector])[0]
        self.__predictedMunsellVector = self.getMunsellValue(predictedValue)

    @property
    def PredictedMunsellVector(self):
        return self.__predictedMunsellVector


if __name__ == "__main__":
    myPredictor = Predictor()
    myPredictor.predictNominalMunsellVector([40, 10, 3])
    print(myPredictor.PredictedMunsellVector)
