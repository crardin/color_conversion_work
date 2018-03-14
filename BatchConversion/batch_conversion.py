import csv
import os
import openpyxl
import pandas as pd
# from sklearn import svm
from sklearn.neighbors import KNeighborsClassifier
from BatchConversion.Munsell import Munsell
from BatchConversion.LABColor import LABColor
from BatchConversion.InputFileHandler import InputFileHandler


class BatchConverter(object):
    # __inputFileName = '../InputData/C&DColors.csv'
    # __inputFileName = '../InputData/TestWithoutAnswers.csv'
    __labTrainingValues = []
    __munsellTrainingValues = []
    __conversionData = {}
    __munsellDataFile = "real_CIELAB.xlsx"
    __inputFileName = ''
    __outputFileName = '../Output/transformedValues.csv'
    __colors = []
    __predictedColors = []
    __inputFileHandler = None

    def __init__(self):
        # self.clf = None
        self.neigh = KNeighborsClassifier(n_neighbors=1)
        self.getTrainingData()
        self.myMunsell = Munsell()
        self.trainData()

    def trainData(self):
        self.neigh.fit(self.__labTrainingValues, self.__munsellTrainingValues)
        # self.clf = svm.SVC()
        # self.clf.fit(self.myMunsell.labValues, self.myMunsell.HValues)
        # self.clf2 = svm.SVC()
        # self.clf2.fit(self.myMunsell.labValues, self.myMunsell.VValues)
        # self.clf3 = svm.SVC()
        # self.clf3.fit(self.myMunsell.labValues, self.myMunsell.CValues)

    def testDataTraining(self):
        # test values
        print(self.clf.predict([[10.63, 12.59, -2.07]]))  # should be 10RP 1/2
        print(self.clf.predict([[10.63, -7.8, 1.3]]))  # should be 7.5G 1/2

    def getTrainingData(self):
        wb = openpyxl.load_workbook(self.__munsellDataFile)
        sheet = wb['data']
        for i in range(2, sheet.max_row + 1):
            munsellValue = self.getMunsellTrainingValue(i, sheet)
            labValue = self.getLabTrainingValue(i, sheet)
            self.__conversionData[labValue] = munsellValue

    def getMunsellTrainingValue(self, i, sheet):
        H = sheet.cell(row=i, column=2).value
        munsellValue = str(H) + " "
        V = sheet.cell(row=i, column=3).value
        munsellValue += str(V) + "/"
        C = sheet.cell(row=i, column=4).value
        munsellValue += str(C)
        self.__munsellTrainingValues.append(str(H + " " + str(V) + "/" + str(C)))
        return munsellValue

    def getLabTrainingValue(self, i, sheet):
        L = sheet.cell(row=i, column=11).value
        labValue = str(L) + " "
        A = sheet.cell(row=i, column=12).value
        labValue += str(A) + " "
        B = sheet.cell(row=i, column=13).value
        labValue += str(B)
        self.__labTrainingValues.append([L, A, B])
        return labValue

    def getInputData(self):
        self.__inputFileHandler.getInputData()
        self.__colors = self.__inputFileHandler.Colors

    def predictData(self):
        self.__predictedColors = []
        for currentColor in self.__colors:
            calculatedValues = currentColor.CalculatedMunsellList
            currentColor = {'colorName': currentColor.colorName, 'L': currentColor.LabList[0], 'A': currentColor.LabList[1],
                            'B': currentColor.LabList[2], 'colorValue': self.neigh.predict([currentColor.LabList])[0],
                            'H1': calculatedValues[0], 'H2': calculatedValues[1], 'V': calculatedValues[2],
                            'C': calculatedValues[3]}
            self.__predictedColors.append(currentColor)

    def outputData(self):
        outputFile = open(self.outputFileName, 'w', newline='')
        outputWriter = csv.writer(outputFile)
        outputWriter.writerow(['Unique #', 'Label', 'L', 'a', 'b', 'Munsell', 'H1', 'H2', 'V', 'C'])
        for color in self.__colors:
            calculatedValues = color.CalculatedMunsellList
            outputWriter.writerow(
                [color.colorIdentifier, color.colorName, color.LabList[0], color.LabList[1], color.LabList[2],
                 self.neigh.predict([color.LabList]), calculatedValues[0], calculatedValues[1], calculatedValues[2],
                 calculatedValues[3]])
        outputFile.close()

    @property
    def conversionData(self):
        return self.__conversionData

    @property
    def predictedColors(self):
        return self.__predictedColors

    @property
    def colors(self):
        self.getInputData()
        return self.__colors

    @property
    def inputFileName(self):
        return self.__inputFileName

    @inputFileName.setter
    def inputFileName(self, value):
        self.__inputFileName = value
        if not self.__inputFileHandler:
            self.__inputFileHandler = InputFileHandler(self.__inputFileName)
        else:
            self.__inputFileName = value
            self.__inputFileHandler.inputFileName = self.__inputFileName

    @property
    def outputFileName(self):
        return self.__outputFileName

    @outputFileName.setter
    def outputFileName(self, value):
        if os.path.isdir(value):
            self.__outputFileName = str(value)


if __name__ == '__main__':
    myBatchConverter = BatchConverter()
    myBatchConverter.inputFileName = '../InputData/test.csv'
    myBatchConverter.getInputData()
    myBatchConverter.predictData()
    for color in myBatchConverter.predictedColors:
        print(color)
    for color in myBatchConverter.colors:
        print(color.CalculatedMunsellList)
