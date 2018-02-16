import openpyxl


class Munsell:
    __munsellDataFile = "../munsell_data/real_CIELAB.xlsx"
    __conversionData = {}
    __labValues = []
    __munsellValues = []
    __H_list = []
    __V_list = []
    __C_list = []
    __L = 0.0
    __A = 0.0
    __B = 0.0
    __H = 0.0
    __V = 0.0
    __C = 0.0

    def __init__(self):
        self.getData()

    def getData(self):
        wb = openpyxl.load_workbook(self.__munsellDataFile)
        sheet = wb['data']
        for i in range(2, sheet.max_row + 1):
            munsellValue = self.getMunsellValue(i, sheet)
            labValue = self.getLabValue(i, sheet)
            self.__conversionData[labValue] = munsellValue

    def getLabValue(self, i, sheet):
        self.__L = sheet.cell(row=i, column=11).value
        labValue = str(self.__L) + " "
        self.__A = sheet.cell(row=i, column=12).value
        labValue += str(self.__A) + " "
        self.__B = sheet.cell(row=i, column=13).value
        labValue += str(self.__B)
        self.__labValues.append([self.__L, self.__A, self.__B])
        return labValue

    def getMunsellValue(self, i, sheet):
        self.__H = sheet.cell(row=i, column=2).value
        munsellValue = str(self.__H) + " "
        self.__V = sheet.cell(row=i, column=3).value
        munsellValue += str(self.__V) + "/"
        self.__C = sheet.cell(row=i, column=4).value
        munsellValue += str(self.__C)
        # self.__munsellValues.append([self.__H, self.__V, self.__C])
        self.__munsellValues.append(str(self.__H) + " " + str(self.__V) + "/" + str(self.__C))
        self.__H_list.append(self.__H)
        self.__V_list.append(float(self.__V))
        self.__C_list.append(float(self.__C))
        return munsellValue

    @property
    def conversionData(self):
        return self.__conversionData

    @property
    def labValues(self):
        return self.__labValues

    @property
    def munsellValues(self):
        return self.__munsellValues

    @property
    def HValues(self):
        return self.__H_list

    @property
    def VValues(self):
        return self.__V_list

    @property
    def CValues(self):
        return self.__C_list
